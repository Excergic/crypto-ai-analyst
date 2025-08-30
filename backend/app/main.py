from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
import logging
import os
from datetime import datetime
from io import BytesIO
import openpyxl
from openpyxl.styles import Font

from .agents.langgraph_workflow import crypto_workflow
from .models.schemas import AnalysisRequest, AnalysisResponse
from .utils.schema_converters import SchemaConverter

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI(
    title="Crypto AI Analyst - LangGraph Architecture",
    description="Multi-agent crypto analysis with proper schema separation",
    version="1.0.0"
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Allow all origins for simplicity
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Create reports directory and mount as static files
os.makedirs("reports", exist_ok=True)
os.makedirs("reports/charts", exist_ok=True)
app.mount("/reports", StaticFiles(directory="reports"), name="reports")

# Serve the HTML frontend
app.mount("/static", StaticFiles(directory="frontend"), name="static")

@app.get("/")
async def root():
    return {
        "message": "Crypto AI Analyst with LangGraph! 🚀",
        "architecture": "dual_schema_approach",
        "features": ["pydantic_validation", "langgraph_workflow", "schema_conversion"]
    }

@app.post("/analyze", response_model=AnalysisResponse)
async def run_analysis(request: AnalysisRequest):
    """Run complete crypto analysis and generate downloadable Excel report"""
    try:
        logger.info(f"🚀 Starting analysis for {request.num_coins} coins...")
        
        # Convert API request to LangGraph state
        initial_state = SchemaConverter.request_to_langgraph_state(request)
        
        # Execute LangGraph workflow
        final_state = await crypto_workflow.ainvoke(initial_state)
        
        # Generate Excel report file
        report_filename = generate_excel_report(final_state)
        
        # Convert back to API response
        response = SchemaConverter.langgraph_state_to_response(final_state)
        
        # Add download links to response
        response.data["report_download_url"] = f"/reports/{report_filename}"
        response.data["report_filename"] = report_filename
        
        logger.info("✅ Analysis completed successfully")
        return response
        
    except Exception as e:
        logger.error(f"❌ Analysis failed: {e}")
        raise HTTPException(status_code=500, detail=f"Analysis error: {str(e)}")

def generate_excel_report(state: dict) -> str:
    """Generate Excel report and save to reports directory"""
    try:
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Crypto Market Analysis"
        
        # Header
        ws.merge_cells("A1:F1")
        ws["A1"] = "Cryptocurrency Market Analysis Report"
        ws["A1"].font = Font(size=14, bold=True)
        
        # Timestamp
        ws["A2"] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # Get data from state
        crypto_data = state.get("validated_crypto_data", [])
        analysis_results = state.get("analysis_results", {})
        
        # Crypto Data Table
        headers = ["ID", "Symbol", "Name", "Current Price", "Market Cap", "24h Change (%)"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
        
        # Fill crypto data
        for row_num, coin in enumerate(crypto_data, 5):
            ws.cell(row=row_num, column=1).value = coin.get("id")
            ws.cell(row=row_num, column=2).value = coin.get("symbol")
            ws.cell(row=row_num, column=3).value = coin.get("name")
            ws.cell(row=row_num, column=4).value = coin.get("current_price")
            ws.cell(row=row_num, column=5).value = coin.get("market_cap")
            ws.cell(row=row_num, column=6).value = coin.get("price_change_percentage_24h")
        
        # Analysis Summary
        summary_start_row = len(crypto_data) + 7
        ws.cell(row=summary_start_row, column=1).value = "Analysis Summary"
        ws.cell(row=summary_start_row, column=1).font = Font(size=12, bold=True)
        
        market_overview = analysis_results.get("market_overview", {})
        row = summary_start_row + 1
        for key, value in market_overview.items():
            ws.cell(row=row, column=1).value = key.replace("_", " ").title()
            ws.cell(row=row, column=2).value = value
            row += 1
        
        # Save to reports directory
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"crypto_analysis_{timestamp}.xlsx"
        filepath = f"reports/{filename}"
        
        wb.save(filepath)
        logger.info(f"📄 Excel report saved: {filename}")
        
        return filename
        
    except Exception as e:
        logger.error(f"❌ Error generating Excel: {e}")
        return ""

@app.get("/health")
async def health_check():
    return {
        "status": "healthy",
        "architecture": "clean_separation",
        "schemas": {
            "pydantic": "API validation",
            "typeddict": "LangGraph workflow"
        }
    }
