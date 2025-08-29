
import logging
from typing import Dict, Any
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from datetime import datetime

logger = logging.getLogger(__name__)

class ReportGenerationAgent:
    def __init__(self):
        self.name = "ReportGenerationAgent"

    def generate_excel_report(self, state: Dict[str, Any]) -> Dict[str, Any]:
        try:
            logger.info("📄 ReportGenerationAgent: Generating Excel report...")
            crypto_data = state.get("validated_crypto_data", [])
            analysis_results = state.get("analysis_results", {})

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Crypto Market Analysis"

            # Header
            ws.merge_cells("A1:F1")
            ws["A1"] = "Cryptocurrency Market Analysis Report"
            ws["A1"].font = Font(size=14, bold=True)

            # Timestamp
            ws["A2"] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

            # Crypto Data Table
            headers = ["ID", "Symbol", "Name", "Current Price", "Market Cap", "24h Change (%)"]
            for col_num, header in enumerate(headers, 1):
                col_letter = get_column_letter(col_num)
                cell = ws[f"{col_letter}4"]
                cell.value = header
                cell.font = Font(bold=True)

            # Fill crypto data
            for row_num, coin in enumerate(crypto_data, 5):
                ws[f"A{row_num}"] = coin.get("id")
                ws[f"B{row_num}"] = coin.get("symbol")
                ws[f"C{row_num}"] = coin.get("name")
                ws[f"D{row_num}"] = coin.get("current_price")
                ws[f"E{row_num}"] = coin.get("market_cap")
                ws[f"F{row_num}"] = coin.get("price_change_percentage_24h")

            # Analysis Summary Section
            ws[f"H4"] = "Analysis Summary"
            ws[f"H4"].font = Font(size=12, bold=True)

            summary = analysis_results.get("market_overview", {})
            row_start = 5
            for key, value in summary.items():
                ws[f"H{row_start}"] = key.replace("_", " ").title()
                ws[f"I{row_start}"] = value
                row_start += 1

            # Save report to file
            report_path = f"reports/crypto_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(report_path)

            # Update state with report path
            state["report_path"] = report_path
            state["report_status"] = "completed"

            logger.info(f"✅ Report generated: {report_path}")
            return state

        except Exception as e:
            logger.error(f"❌ ReportGenerationAgent error: {e}")
            state["report_status"] = "failed"
            state["report_error"] = str(e)
            return state


# Create agent instance
report_generation_agent = ReportGenerationAgent()
